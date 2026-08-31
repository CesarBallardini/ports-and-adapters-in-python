"""The XLSX spreadsheet adapter: reader and writer over ``openpyxl``.

The harder of the two adapters, and the reason the port insists on what it does. A workbook
knows its cells' types, so this adapter has to *give that knowledge up*: a cell containing the
number 8 must arrive as ``'8'``, exactly as the CSV adapter would deliver it, or the two
implementations diverge and the rules above the port start behaving differently depending on
what the registrar happened to upload (ADR-0008).

Every rendering choice below is therefore a decision about matching CSV, not about being
faithful to Excel.
"""

from __future__ import annotations

import datetime as dt
from io import BytesIO
from typing import Any, cast
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from academy.application.errors import MalformedSpreadsheetError


class XlsxSpreadsheetReader:
    """Parses XLSX bytes into rows of strings.

    Satisfies :class:`~academy.application.ports.outbound.spreadsheet.SpreadsheetReader`.
    """

    def read_rows(self, data: bytes) -> list[dict[str, str]]:
        """Parse the first worksheet of ``data`` into one dictionary per data row.

        The *first* sheet, because the port promises one table per file and a workbook with
        several is a file whose author meant something this system has no way to ask about.

        Returns:
            One dictionary per row after the header, in sheet order, padded and truncated to
            the header's width. A workbook with no rows, or only a header, yields an empty
            list.

        Raises:
            MalformedSpreadsheetError: If the bytes are not a readable workbook. ``openpyxl``
                alone raises ``InvalidFileException``, ``BadZipFile``, ``KeyError``,
                ``OSError`` and ``ValueError`` for that one condition; all of them are
                normalised here, which is the whole reason this adapter exists.
        """
        workbook = self._load(data)
        try:
            sheet = workbook.worksheets[0]
            rows = [[_cell(value) for value in row] for row in sheet.iter_rows(values_only=True)]
        finally:
            workbook.close()

        header_and_body = [row for row in rows if any(cell.strip() for cell in row)]
        if not header_and_body:
            return []

        header, *body = header_and_body
        return [_row(_named(header), cells) for cells in body]

    @staticmethod
    def _load(data: bytes) -> Workbook:
        """Open the workbook, or normalise whichever failure ``openpyxl`` chose.

        Raises:
            MalformedSpreadsheetError: For every one of them.
        """
        try:
            return load_workbook(BytesIO(data), read_only=True, data_only=True)
        except (InvalidFileException, BadZipFile, KeyError, OSError, ValueError, TypeError) as error:
            raise MalformedSpreadsheetError(f'not a readable workbook: {error}') from error


class XlsxSpreadsheetWriter:
    """Renders rows of strings as XLSX bytes.

    Satisfies :class:`~academy.application.ports.outbound.spreadsheet.SpreadsheetWriter`.
    """

    def write_sheet(
        self,
        headers: list[str],
        rows: list[list[str]],
        *,
        sheet_name: str = 'Sheet1',
    ) -> bytes:
        """Render a header row and data rows as a workbook.

        Every cell is written as text. A template whose ids Excel decided were numbers, and
        then rendered as ``1.23457E+11``, is a template that comes back unusable.

        Raises:
            ValueError: If any row's length differs from ``headers``.
        """
        _require_rectangular(headers, rows)

        workbook = Workbook()
        try:
            sheet = workbook.active
            if sheet is None:  # pragma: no cover - a new Workbook always has an active sheet
                sheet = workbook.create_sheet()
            sheet.title = sheet_name
            if headers:
                sheet.append(headers)
            for row in rows:
                sheet.append(row)

            buffer = BytesIO()
            workbook.save(buffer)
        finally:
            workbook.close()
        return buffer.getvalue()


def _cell(value: Any) -> str:  # noqa: ANN401 - openpyxl hands back whatever the cell holds
    """Render one cell the way the CSV adapter would have delivered it.

    Each case is a match against CSV rather than a rendering preference:

    * an empty cell is ``''``, because that is what a missing field looks like in a CSV row;
    * a whole number is ``'8'`` and not ``'8.0'`` -- Excel stores integers as floats, and the
      port's own example says a cell containing 8 must arrive as ``'8'``;
    * a date is ISO-8601, the only rendering the domain's value objects parse and the only one
      that does not depend on a locale. Excel has no date-only type -- it stores a serial
      number that ``openpyxl`` hands back as a ``datetime`` at midnight -- so a timestamp of
      exactly midnight is rendered as a plain date. The cost is that a genuine midnight
      timestamp loses its time; the benefit is that every birth date in every uploaded sheet
      arrives as ``'2011-05-01'`` rather than ``'2011-05-01T00:00:00'``, which is what CSV
      would have carried and what the value objects parse;
    * a boolean is ``'TRUE'``/``'FALSE'``, which is what a spreadsheet shows the person who
      typed it and therefore what they would have exported to CSV.
    """
    match value:
        case None:
            return ''
        case bool():
            return 'TRUE' if value else 'FALSE'
        case dt.datetime() if value.time() == dt.time.min:
            return value.date().isoformat()
        case dt.datetime():
            return value.isoformat()
        case dt.date():
            return value.isoformat()
        case float() if value.is_integer():
            return str(int(value))
        case _:
            return str(cast(object, value))


def _named(header: list[str]) -> list[str]:
    """Drop the trailing columns that have no name.

    Load-bearing here in a way it is not for CSV. ``openpyxl`` pads every row to the sheet's
    widest, so a data row with one cell more than the header comes back having *extended* the
    header with an unnamed column -- and that cell would then survive under the key ``''``,
    where the port says cells beyond the header are dropped. Without this, the two adapters
    disagree about a ragged file, which is precisely the file a human produces.

    Trailing only: an unnamed column between two named ones is a stranger thing, and dropping
    it would silently shift every cell after it.
    """
    while header and not header[-1].strip():
        header = header[:-1]
    return header


def _row(header: list[str], cells: list[str]) -> dict[str, str]:
    """Pair one row's cells with the header, padding short rows and dropping extra cells."""
    padded = cells[: len(header)] + [''] * (len(header) - len(cells))
    return dict(zip(header, padded, strict=True))


def _require_rectangular(headers: list[str], rows: list[list[str]]) -> None:
    """Refuse rows that do not match the header's width.

    Raises:
        ValueError: Naming the offending row.
    """
    for number, row in enumerate(rows, start=1):
        if len(row) != len(headers):
            raise ValueError(f'row {number} has {len(row)} cells, but there are {len(headers)} headers')
