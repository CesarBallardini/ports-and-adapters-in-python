"""The CLI over a real, migrated database, driven in process.

Everything below the ``main()`` call is production code: the real composition root, the real
SQLAlchemy repositories, the real schema built by Alembic from empty (ADR-0006). Only argv, the
two streams and the environment are supplied, which is exactly what a shell supplies.

This is where claim 3 stops being a slogan. The fixtures seed through the *repositories*, and the
commands then reach the same ``ImportService``, ``GradeManagement`` and ``StudentRecords`` the
acceptance suite drives over the memory backend -- with no HTTP anywhere in the process and no
line of the application aware that argv exists.

Not the e2e tier: that one spawns the real interpreter and checks the status a shell sees. This
one can assert on the database afterwards, which the other cannot.
"""

from __future__ import annotations

import asyncio
import io
import json
from collections.abc import AsyncIterator
from datetime import date
from pathlib import Path
from typing import Any
from uuid import UUID

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from academy.adapters.inbound.cli import ExitCode, main
from academy.adapters.outbound.persistence.sqlalchemy.repositories import (
    SqlAlchemyAcademicHistoryRepository,
    SqlAlchemyConfigurationRepository,
    SqlAlchemyGuardianshipRepository,
    SqlAlchemyPersonRepository,
    SqlAlchemySectionRepository,
)
from academy.adapters.outbound.persistence.sqlalchemy.session import (
    create_engine,
    create_session_factory,
    migrate_to_head,
)
from academy.adapters.outbound.persistence.sqlalchemy.unit_of_work import SqlAlchemyUnitOfWork
from academy.application.jobs import JobStatus
from academy.config.settings import (
    ENV_DATABASE_URL,
    ENV_IMPORT_INLINE_THRESHOLD,
    ENV_PERSISTENCE,
    ENV_UPLOAD_DIRECTORY,
)
from academy.domain.academics.course_section import CourseSection
from academy.domain.academics.term import Term
from academy.domain.grades.academic_history import AcademicHistory
from academy.domain.grades.grade import Grade
from academy.domain.grades.grade_entry import GradeEntry
from academy.domain.guardianship.guardianship import Guardianship
from academy.domain.people.age_of_majority import AgeOfMajority
from academy.domain.people.email import Email
from academy.domain.people.person import Person
from academy.domain.people.personal_data import PersonalData
from academy.domain.people.role import Role
from academy.domain.shared.ids import GuardianshipId, PersonId, SectionId, SubjectId

pytestmark = pytest.mark.integration

TEACHER = PersonId(UUID(int=1))
STUDENT = PersonId(UUID(int=2))
OTHER_STUDENT = PersonId(UUID(int=3))
GUARDIAN = PersonId(UUID(int=4))
WARD = PersonId(UUID(int=5))
SECTION = SectionId(UUID(int=10))
MATHEMATICS = SubjectId(UUID(int=20))
TERM = Term(2026, 1)

TEACHER_EMAIL = 'tess@academy.test'
STUDENT_EMAIL = 'sam@academy.test'
GUARDIAN_EMAIL = 'gale@academy.test'


class Deployment:
    """One seeded database, plus the environment that points the CLI at it."""

    def __init__(self, url: str, uploads: Path) -> None:
        """Record where the database and the payload directory are."""
        self.environ = {
            ENV_PERSISTENCE: 'sqlalchemy',
            ENV_DATABASE_URL: url,
            ENV_UPLOAD_DIRECTORY: str(uploads),
        }

    def run(self, *argv: str) -> tuple[int, str, str]:
        """Run one command line the way a shell would.

        Returns:
            The exit status, what went to stdout and what went to stderr -- separately, because
            a shell can pipe one without catching the other and a test should be able to tell.
        """
        out, err = io.StringIO(), io.StringIO()
        code = main(argv, out=out, err=err, environ=self.environ)
        return code, out.getvalue(), err.getvalue()

    def json(self, *argv: str) -> tuple[int, dict[str, Any]]:
        """Run a command with ``--json`` and parse what it printed.

        ``Any`` is the honest type of a parsed JSON document and one of the few places it is
        (Rule 2). The shape is checked by the assertion that follows in each test, which is a
        stronger statement than any annotation here could make: narrowing every access would
        triple the length of these tests without checking anything the assertion does not.
        """
        code, out, _ = self.run(*argv, '--json')
        return code, json.loads(out)


@pytest.fixture
async def deployment(tmp_path: Path) -> AsyncIterator[Deployment]:
    """A migrated SQLite database with enough people in it to exercise every command.

    Migrated, never ``create_all`` (ADR-0006): the schema the CLI reads is the schema a
    deployment gets. The migration runs in a worker thread because Alembic drives its own event
    loop, and ``asyncio.run`` inside a running one is an error.
    """
    url = f'sqlite+aiosqlite:///{(tmp_path / "academy.db").as_posix()}'
    await asyncio.to_thread(migrate_to_head, url)

    engine = create_engine(url)
    try:
        async with create_session_factory(engine)() as session:
            await _seed(session)
        yield Deployment(url, tmp_path / 'uploads')
    finally:
        # Windows will not delete the temporary directory while a handle is open on the file.
        await engine.dispose()


async def _seed(session: AsyncSession) -> None:
    """Put a teacher, two students, a guardian and a ward into the database."""
    people = SqlAlchemyPersonRepository(session)
    sections = SqlAlchemySectionRepository(session)
    histories = SqlAlchemyAcademicHistoryRepository(session)
    guardianships = SqlAlchemyGuardianshipRepository(session)
    configuration = SqlAlchemyConfigurationRepository(session)
    unit_of_work = SqlAlchemyUnitOfWork(session)

    async with unit_of_work:
        await people.add(_person(TEACHER, TEACHER_EMAIL, 'Tess Teacher', date(1980, 1, 1), Role.TEACHER))
        await people.add(_person(STUDENT, STUDENT_EMAIL, 'Sam Student', date(2000, 1, 1), Role.STUDENT))
        await people.add(_person(OTHER_STUDENT, 'oona@academy.test', 'Oona Other', date(2000, 1, 1), Role.STUDENT))
        await people.add(_person(GUARDIAN, GUARDIAN_EMAIL, 'Gale Guardian', date(1980, 1, 1)))
        await people.add(_person(WARD, 'wren@academy.test', 'Wren Ward', date(2015, 1, 1), Role.STUDENT))

        section = CourseSection(id=SECTION, subject_id=MATHEMATICS, term=TERM, teacher_id=TEACHER)
        section.enroll(STUDENT)
        await sections.add(section)

        history = AcademicHistory(student_id=STUDENT)
        history.record(GradeEntry(subject_id=MATHEMATICS, term=TERM, grade=Grade(7), source_section_id=SECTION))
        await histories.add(history)

        await guardianships.add(Guardianship(id=GuardianshipId(UUID(int=30)), guardian_id=GUARDIAN, ward_id=WARD))
        await configuration.set_age_of_majority(AgeOfMajority(18))
        await unit_of_work.commit()


def _person(person_id: PersonId, email: str, name: str, born: date, *roles: Role) -> Person:
    return Person(
        id=person_id,
        email=Email(email),
        personal=PersonalData(full_name=name, birth_date=born),
        roles=set(roles),
    )


def _filled_template(deployment: Deployment, path: Path, *, grade: int) -> Path:
    """Download an XLSX template through the CLI and fill in the one enrolled student's grade.

    The round trip a registrar actually makes, and the only way to produce a *valid* XLSX payload
    without hand-building a workbook the CLI never would have written.
    """
    deployment.run(
        'import', 'template', '--format', 'xlsx', '--section', str(SECTION), '--output', str(path),
        '--as', TEACHER_EMAIL,
    )  # fmt: skip

    workbook = load_workbook(path)
    sheet = workbook.active
    assert sheet is not None
    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    sheet.cell(row=2, column=headers.index('grade') + 1, value=grade)
    workbook.save(path)
    return path


# --------------------------------------------------------------------------------------
# The actor: asserted by --as, and never stale (ADR-0020)
# --------------------------------------------------------------------------------------


@pytest.mark.integration
def test_an_address_that_names_nobody_is_not_found_rather_than_forbidden(deployment: Deployment) -> None:
    """A typo should send its operator to the address bar, not into the policy."""
    code, out, err = deployment.run('records', 'wards', '--as', 'dana@exmaple.test')

    assert code == ExitCode.NOT_FOUND
    assert out == ''
    assert 'exmaple' in err


@pytest.mark.integration
def test_the_actor_carries_the_roles_the_person_holds_right_now(deployment: Deployment) -> None:
    """An actor built from an id alone has no roles, so it is a different actor, not a smaller one.

    The teacher can read their own section only because ``--as`` resolved to a person whose
    ``TEACHER`` role was read from the record on this invocation.
    """
    code, out, _ = deployment.run('grades', 'list', str(SECTION), '--as', TEACHER_EMAIL)

    assert code == ExitCode.OK
    assert 'Sam Student' in out


@pytest.mark.integration
def test_asserting_an_identity_does_not_grant_what_that_identity_lacks(deployment: Deployment) -> None:
    """``--as`` says who you claim to be; the policy still says what that person may do."""
    code, _, err = deployment.run('grades', 'record', str(SECTION), str(STUDENT), '9', '--as', STUDENT_EMAIL)

    assert code == ExitCode.FORBIDDEN
    assert err.startswith('forbidden: ')


# --------------------------------------------------------------------------------------
# Grades and records
# --------------------------------------------------------------------------------------


@pytest.mark.integration
def test_a_teacher_records_a_grade_and_it_is_in_the_database_afterwards(deployment: Deployment) -> None:
    code, data = deployment.json('grades', 'record', str(SECTION), str(STUDENT), '9', '--as', TEACHER_EMAIL)

    assert code == ExitCode.OK
    assert data['recorded_grade'] == 9
    assert data['best_grade'] == 9

    # Read it back through a *different* invocation, which is the only way to know it committed.
    _, transcript = deployment.json('records', 'show', str(STUDENT), '--as', STUDENT_EMAIL)
    assert [entry['grade'] for entry in transcript['entries']] == [7, 9]


@pytest.mark.integration
def test_recording_a_worse_grade_reports_that_the_standing_did_not_move(deployment: Deployment) -> None:
    """The question a teacher actually has, which an acknowledgement would not answer."""
    code, data = deployment.json('grades', 'record', str(SECTION), str(STUDENT), '4', '--as', TEACHER_EMAIL)

    assert code == ExitCode.OK
    assert (data['recorded_grade'], data['best_grade'], data['passed']) == (4, 7, True)


@pytest.mark.integration
def test_a_grade_outside_the_scale_is_refused_by_the_domain_with_a_validation_status(deployment: Deployment) -> None:
    """``11`` parses as an integer and is stopped where it should be, not by the parser."""
    code, _, err = deployment.run('grades', 'record', str(SECTION), str(STUDENT), '11', '--as', TEACHER_EMAIL)

    assert code == ExitCode.VALIDATION
    assert err.startswith('validation: ')


@pytest.mark.integration
def test_a_student_reads_their_own_transcript(deployment: Deployment) -> None:
    code, data = deployment.json('records', 'show', str(STUDENT), '--as', STUDENT_EMAIL)

    assert code == ExitCode.OK
    assert data['student_id'] == str(STUDENT)
    assert data['standings'] == [
        {'subject_id': str(MATHEMATICS), 'best_grade': 7, 'passed': True, 'attempts': 1},
    ]


@pytest.mark.integration
def test_a_student_may_not_read_another_students_transcript(deployment: Deployment) -> None:
    code, _, _ = deployment.run('records', 'show', str(STUDENT), '--as', 'oona@academy.test')

    assert code == ExitCode.FORBIDDEN


@pytest.mark.integration
def test_a_guardian_lists_the_wards_currently_in_their_care(deployment: Deployment) -> None:
    code, data = deployment.json('records', 'wards', '--as', GUARDIAN_EMAIL)

    assert code == ExitCode.OK
    assert [ward['full_name'] for ward in data['wards']] == ['Wren Ward']


@pytest.mark.integration
def test_someone_with_no_wards_gets_an_empty_list_rather_than_a_refusal(deployment: Deployment) -> None:
    code, data = deployment.json('records', 'wards', '--as', TEACHER_EMAIL)

    assert code == ExitCode.OK
    assert data['wards'] == []


# --------------------------------------------------------------------------------------
# Import: the fourth driver over the same use case
# --------------------------------------------------------------------------------------


@pytest.mark.integration
def test_a_template_comes_back_prefilled_with_the_enrolled_students(deployment: Deployment, tmp_path: Path) -> None:
    """UC-36: the round trip is a workflow, not a blank form."""
    template = tmp_path / 'template.csv'

    code, _, _ = deployment.run(
        'import', 'template', '--format', 'csv', '--section', str(SECTION), '--output', str(template),
        '--as', TEACHER_EMAIL,
    )  # fmt: skip

    assert code == ExitCode.OK
    assert STUDENT_EMAIL in template.read_text(encoding='utf-8')


@pytest.mark.integration
def test_a_dry_run_reports_what_would_happen_and_changes_nothing(deployment: Deployment, tmp_path: Path) -> None:
    """The most useful thing on the whole import surface, and the reason exit 10 exists."""
    sheet = tmp_path / 'grades.csv'
    sheet.write_text(f'student_email,grade\n{STUDENT_EMAIL},10\n', encoding='utf-8')

    code, data = deployment.json(
        'import', 'run', str(sheet), '--section', str(SECTION), '--dry-run', '--as', TEACHER_EMAIL
    )

    assert code == ExitCode.OK
    assert data['dry_run'] is True
    assert data['updated'] + data['created'] == 1

    _, transcript = deployment.json('records', 'show', str(STUDENT), '--as', STUDENT_EMAIL)
    assert [entry['grade'] for entry in transcript['entries']] == [7]


@pytest.mark.integration
def test_a_real_run_writes_the_grades(deployment: Deployment, tmp_path: Path) -> None:
    sheet = tmp_path / 'grades.csv'
    sheet.write_text(f'student_email,grade\n{STUDENT_EMAIL},10\n', encoding='utf-8')

    code, _, _ = deployment.run('import', 'run', str(sheet), '--section', str(SECTION), '--as', TEACHER_EMAIL)

    assert code == ExitCode.OK
    _, transcript = deployment.json('records', 'show', str(STUDENT), '--as', STUDENT_EMAIL)
    assert sorted(entry['grade'] for entry in transcript['entries']) == [7, 10]


@pytest.mark.integration
def test_a_rejected_row_is_reported_and_exits_ten_without_stopping_the_file(
    deployment: Deployment, tmp_path: Path
) -> None:
    """Partial success is the point: the good row lands, the bad one comes back with its line."""
    sheet = tmp_path / 'grades.csv'
    sheet.write_text(f'student_email,grade\n{STUDENT_EMAIL},10\nnobody@academy.test,8\n', encoding='utf-8')

    code, data = deployment.json('import', 'run', str(sheet), '--section', str(SECTION), '--as', TEACHER_EMAIL)

    assert code == ExitCode.IMPORT_INCOMPLETE
    assert data['errors'] != []
    assert data['created'] + data['updated'] == 1


@pytest.mark.integration
def test_a_file_that_is_not_a_spreadsheet_is_a_validation_failure(deployment: Deployment, tmp_path: Path) -> None:
    """One error type for every unreadable file, whichever adapter met it (ADR-0008)."""
    sheet = tmp_path / 'grades.xlsx'
    sheet.write_bytes(b'this is not a workbook')

    code, _, err = deployment.run('import', 'run', str(sheet), '--section', str(SECTION), '--as', TEACHER_EMAIL)

    assert code == ExitCode.VALIDATION
    assert err.startswith('validation: ')


@pytest.mark.integration
def test_an_xlsx_template_round_trips_back_through_the_importer(deployment: Deployment, tmp_path: Path) -> None:
    """Download, fill in, upload: the two spreadsheet adapters meeting through the CLI."""
    template = _filled_template(deployment, tmp_path / 'template.xlsx', grade=6)

    code, data = deployment.json('import', 'run', str(template), '--section', str(SECTION), '--as', TEACHER_EMAIL)

    assert code == ExitCode.OK
    assert data['created'] + data['updated'] == 1


@pytest.mark.integration
def test_an_import_above_the_threshold_is_queued_and_still_reports_its_outcome(
    deployment: Deployment, tmp_path: Path
) -> None:
    """The queued path, reached by lowering the threshold rather than by a flag (ADR-0009).

    The deployment wires an inline queue, so the job has already run by the time the command
    reports it -- which is why the handler re-reads the job rather than reporting the object
    ``submit`` handed back.
    """
    deployment.environ[ENV_IMPORT_INLINE_THRESHOLD] = '1'
    sheet = tmp_path / 'grades.csv'
    sheet.write_text(f'student_email,grade\n{STUDENT_EMAIL},10\n', encoding='utf-8')

    code, data = deployment.json('import', 'run', str(sheet), '--section', str(SECTION), '--as', TEACHER_EMAIL)

    assert code == ExitCode.OK
    assert data['status'] == JobStatus.DONE.value
    assert data['result'] is not None


@pytest.mark.integration
def test_a_queued_xlsx_is_read_by_the_xlsx_adapter_and_not_by_the_default_one(
    deployment: Deployment, tmp_path: Path
) -> None:
    """The regression the CLI found: a job carries no MIME type, so the key carries the format.

    Before the storage key gained the payload's extension, every queued XLSX was read back with
    the default CSV reader and failed as malformed.
    """
    template = _filled_template(deployment, tmp_path / 'sheet.xlsx', grade=6)
    deployment.environ[ENV_IMPORT_INLINE_THRESHOLD] = '1'

    code, data = deployment.json('import', 'run', str(template), '--section', str(SECTION), '--as', TEACHER_EMAIL)

    assert code == ExitCode.OK
    assert data['status'] == JobStatus.DONE.value
    assert data['failure_reason'] is None


@pytest.mark.integration
def test_a_job_can_be_read_back_by_id(deployment: Deployment, tmp_path: Path) -> None:
    deployment.environ[ENV_IMPORT_INLINE_THRESHOLD] = '1'
    sheet = tmp_path / 'grades.csv'
    sheet.write_text(f'student_email,grade\n{STUDENT_EMAIL},10\n', encoding='utf-8')
    _, submitted = deployment.json('import', 'run', str(sheet), '--section', str(SECTION), '--as', TEACHER_EMAIL)

    code, data = deployment.json('import', 'job', str(submitted['job_id']), '--as', TEACHER_EMAIL)

    assert code == ExitCode.OK
    assert data['job_id'] == submitted['job_id']


@pytest.mark.integration
def test_an_unknown_job_id_is_not_found(deployment: Deployment) -> None:
    code, _, _ = deployment.run('import', 'job', str(UUID(int=999)), '--as', TEACHER_EMAIL)

    assert code == ExitCode.NOT_FOUND


# --------------------------------------------------------------------------------------
# Configuration
# --------------------------------------------------------------------------------------


@pytest.mark.integration
def test_config_show_reports_the_environment_it_was_given(deployment: Deployment) -> None:
    code, data = deployment.json('config', 'show')

    assert code == ExitCode.OK
    assert data['persistence'] == 'sqlalchemy'
    assert data['database_url'] == deployment.environ[ENV_DATABASE_URL]


@pytest.mark.integration
def test_config_show_needs_neither_a_database_nor_an_actor(tmp_path: Path) -> None:
    """It reaches no use case, so it must work on a deployment whose database does not exist yet."""
    environ = {
        ENV_PERSISTENCE: 'sqlalchemy',
        ENV_DATABASE_URL: f'sqlite+aiosqlite:///{(tmp_path / "missing.db").as_posix()}',
    }
    out = io.StringIO()

    assert main(('config', 'show'), out=out, err=io.StringIO(), environ=environ) == ExitCode.OK
    assert 'persistence = sqlalchemy' in out.getvalue()


@pytest.mark.integration
def test_an_unreadable_setting_refuses_to_start_rather_than_failing_later(tmp_path: Path) -> None:
    """Exit 9, before a use case is reached: there is no HTTP status for this and no request to fail."""
    del tmp_path
    err = io.StringIO()

    code = main(('config', 'show'), out=io.StringIO(), err=err, environ={ENV_PERSISTENCE: 'postgres'})

    assert code == ExitCode.CONFIGURATION
    assert 'memory, sqlalchemy' in err.getvalue()


@pytest.mark.integration
def test_results_go_to_stdout_and_failures_go_to_stderr(deployment: Deployment) -> None:
    """So a shell can pipe one without catching the other."""
    ok_code, ok_out, ok_err = deployment.run('records', 'wards', '--as', GUARDIAN_EMAIL)
    bad_code, bad_out, bad_err = deployment.run('records', 'show', str(STUDENT), '--as', 'oona@academy.test')

    assert (ok_code, ok_err) == (ExitCode.OK, '')
    assert ok_out != ''
    assert (bad_code, bad_out) == (ExitCode.FORBIDDEN, '')
    assert bad_err != ''
